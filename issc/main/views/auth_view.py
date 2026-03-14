from django.http import HttpResponse,JsonResponse
from django.template import loader
from django.shortcuts import render, redirect, get_object_or_404

from django.contrib.auth import authenticate
import secrets
import string
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout as auth_logout
from django.contrib.auth.decorators import login_required

from django.core.files.storage import FileSystemStorage
from django.core.paginator import Paginator

from ..models import AccountRegistration, IncidentReport, VehicleRegistration, SystemConfig

from .utils import paginate

from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.views import PasswordResetView
from django.contrib import messages
from django.core.mail import send_mail, BadHeaderError, get_connection
from django.conf import settings
from django.contrib.auth.hashers import check_password
from django.utils import timezone
from datetime import timedelta
import re

import pandas as pd
from datetime import datetime


PASSWORD_POLICY_MESSAGE = (
    'Password must be at least 8 characters and include at least 1 uppercase letter, '
    '1 lowercase letter, 1 number, and 1 special character.'
)


def sync_admin_flags(user_obj, privilege=None):
    """Keep Django admin flags aligned with custom privilege."""
    target_privilege = (privilege if privilege is not None else getattr(user_obj, 'privilege', '')) or ''
    updates = []

    if target_privilege == 'admin' and not user_obj.is_staff:
        user_obj.is_staff = True
        updates.append('is_staff')
    elif target_privilege != 'admin' and user_obj.is_staff and not user_obj.is_superuser:
        user_obj.is_staff = False
        updates.append('is_staff')

    if updates:
        user_obj.save(update_fields=updates)


def validate_password_strength(password):
    if len(password or '') < 8:
        return False
    if not re.search(r'[A-Z]', password):
        return False
    if not re.search(r'[a-z]', password):
        return False
    if not re.search(r'\d', password):
        return False
    if not re.search(r'[^A-Za-z0-9]', password):
        return False
    return True


def generate_login_otp():
    return ''.join(secrets.choice(string.digits) for _ in range(6))


def send_login_otp_email(user, otp_code):
    subject = 'ISSC Login Verification Code'
    message = (
        f'Hello {user.first_name},\n\n'
        f'Your ISSC one-time verification code is: {otp_code}\n\n'
        'This code will expire in 10 minutes.\n\n'
        'If you did not request this login, please contact your administrator immediately.'
    )
    from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', settings.EMAIL_HOST_USER)
    send_mail(subject, message, from_email, [user.email], fail_silently=False)


def login(request):
    template = loader.get_template('login.html')
    context = {
        'show_otp_modal': False,
        'show_force_password_modal': False,
        'password_policy_message': PASSWORD_POLICY_MESSAGE,
    }
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.session.get('otp_pending_user_id'):
        context['show_otp_modal'] = True
    if request.session.get('force_password_user_id'):
        context['show_force_password_modal'] = True

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()

        if action == 'verify_otp':
            pending_user_id = request.session.get('otp_pending_user_id')
            if not pending_user_id:
                context['error'] = 'Your verification session expired. Please login again.'
                return HttpResponse(template.render(context, request))

            user = AccountRegistration.objects.filter(id=pending_user_id).first()
            if not user:
                request.session.pop('otp_pending_user_id', None)
                context['error'] = 'Account not found. Please login again.'
                return HttpResponse(template.render(context, request))

            now = timezone.now()
            if user.login_otp_locked_until and now < user.login_otp_locked_until:
                remaining_seconds = int((user.login_otp_locked_until - now).total_seconds())
                remaining_minutes = max(1, (remaining_seconds + 59) // 60)
                context['error'] = f'Too many failed attempts. Try again in {remaining_minutes} minute(s).'
                context['show_otp_modal'] = True
                return HttpResponse(template.render(context, request))

            otp_input = request.POST.get('otp_code', '').strip()
            if not otp_input:
                context['error'] = 'Please enter the OTP code.'
                context['show_otp_modal'] = True
                return HttpResponse(template.render(context, request))

            if not user.login_otp_code or not user.login_otp_expires_at or now > user.login_otp_expires_at:
                user.login_otp_code = ''
                user.login_otp_expires_at = None
                user.login_otp_failed_attempts = 0
                user.save(update_fields=['login_otp_code', 'login_otp_expires_at', 'login_otp_failed_attempts'])
                context['error'] = 'OTP expired. Please login again to request a new code.'
                request.session.pop('otp_pending_user_id', None)
                return HttpResponse(template.render(context, request))

            if otp_input != user.login_otp_code:
                user.login_otp_failed_attempts += 1
                if user.login_otp_failed_attempts >= 5:
                    user.login_otp_locked_until = now + timedelta(minutes=5)
                    user.login_otp_failed_attempts = 0
                    user.login_otp_code = ''
                    user.login_otp_expires_at = None
                    user.save(update_fields=['login_otp_failed_attempts', 'login_otp_locked_until', 'login_otp_code', 'login_otp_expires_at'])
                    context['error'] = 'Too many failed OTP attempts. Try again after 5 minutes.'
                    request.session.pop('otp_pending_user_id', None)
                    return HttpResponse(template.render(context, request))

                remaining_attempts = 5 - user.login_otp_failed_attempts
                user.save(update_fields=['login_otp_failed_attempts'])
                context['error'] = f'Invalid OTP. {remaining_attempts} attempt(s) remaining.'
                context['show_otp_modal'] = True
                return HttpResponse(template.render(context, request))

            user.login_otp_code = ''
            user.login_otp_expires_at = None
            user.login_otp_failed_attempts = 0
            user.login_otp_locked_until = None
            user.save(update_fields=['login_otp_code', 'login_otp_expires_at', 'login_otp_failed_attempts', 'login_otp_locked_until'])

            request.session.pop('otp_pending_user_id', None)
            request.session['force_password_user_id'] = user.id
            context['show_force_password_modal'] = True
            context['show_otp_modal'] = False
            return HttpResponse(template.render(context, request))

        if action == 'set_first_password':
            force_user_id = request.session.get('force_password_user_id')
            if not force_user_id:
                context['error'] = 'Session expired. Please login again.'
                return HttpResponse(template.render(context, request))

            user = AccountRegistration.objects.filter(id=force_user_id).first()
            if not user:
                request.session.pop('force_password_user_id', None)
                context['error'] = 'Account not found. Please login again.'
                return HttpResponse(template.render(context, request))

            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')

            if new_password != confirm_password:
                context['error'] = 'Passwords do not match.'
                context['show_force_password_modal'] = True
                return HttpResponse(template.render(context, request))

            if not validate_password_strength(new_password):
                context['error'] = PASSWORD_POLICY_MESSAGE
                context['show_force_password_modal'] = True
                return HttpResponse(template.render(context, request))

            user.set_password(new_password)
            user.must_change_password = False
            user.login_otp_code = ''
            user.login_otp_expires_at = None
            user.login_otp_failed_attempts = 0
            user.login_otp_locked_until = None
            user.save()

            request.session.pop('force_password_user_id', None)
            auth_login(request, user)
            return redirect('dashboard')

        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            sync_admin_flags(user)

            if user.must_change_password:
                now = timezone.now()
                if user.login_otp_locked_until and now < user.login_otp_locked_until:
                    remaining_seconds = int((user.login_otp_locked_until - now).total_seconds())
                    remaining_minutes = max(1, (remaining_seconds + 59) // 60)
                    context['error'] = f'Too many failed OTP attempts. Try again in {remaining_minutes} minute(s).'
                    return HttpResponse(template.render(context, request))

                otp_code = generate_login_otp()
                user.login_otp_code = otp_code
                user.login_otp_expires_at = now + timedelta(minutes=10)
                user.login_otp_failed_attempts = 0
                user.save(update_fields=['login_otp_code', 'login_otp_expires_at', 'login_otp_failed_attempts'])

                try:
                    send_login_otp_email(user, otp_code)
                except Exception as e:
                    context['error'] = f'Failed to send OTP email: {e}'
                    return HttpResponse(template.render(context, request))

                request.session['otp_pending_user_id'] = user.id
                context['show_otp_modal'] = True
                context['otp_email'] = user.email
                return HttpResponse(template.render(context, request))

            auth_login(request, user)
            return redirect('dashboard')

        context['error'] = 'Invalid username or password.'

    return HttpResponse(template.render(context, request))



# @login_required(login_url='/login/')
def signup(request):
    user = AccountRegistration.objects.filter(username=request.user).values()

    users_list = AccountRegistration.objects.all().order_by('-date_joined')
    # users = paginate(users_list,request)


    template = loader.get_template('signup.html')
    context = {
        'user_role': user[0]['privilege'],
        'user_data':user[0],
        'users':users_list
    }

    if request.method == 'POST':
        username = request.POST.get('username')
        if 'delete' in request.POST:
            # Handle delete action
            user = AccountRegistration.objects.get(username=username)
            user.delete()
            return redirect('signup')  # Redirect to the same page after deleting

        if 'update' in request.POST:
            # Handle update action (you can collect and update data as needed)
            user = AccountRegistration.objects.get(username=username)
            new_email = request.POST.get('email')
            new_id_number = request.POST.get('id_number')
            
            # Check for duplicate email (excluding current user)
            if AccountRegistration.objects.filter(email=new_email).exclude(username=username).exists():
                messages.error(request, f"An account with the email '{new_email}' already exists. Please use a different email address.")
                return redirect('signup')
            
            # Check for duplicate ID number (excluding current user)
            if AccountRegistration.objects.filter(id_number=new_id_number).exclude(username=username).exists():
                messages.error(request, f"An account with the ID number '{new_id_number}' already exists. Please use a different ID number.")
                return redirect('signup')
            
            try:
                user.first_name = request.POST.get('first_name')
                user.middle_name = request.POST.get('middle_name')
                user.last_name = request.POST.get('last_name')
                user.email = new_email
                user.id_number = new_id_number
                user.contact_number = request.POST.get('contact_number')
                user.gender = request.POST.get('gender')
                user.department = request.POST.get('department')
                user.privilege = request.POST.get('privilege')
                user.status = request.POST.get('status')
                user.save()
                sync_admin_flags(user)
                messages.success(request, 'Account updated successfully!')
                print('')
                return redirect('signup')  # Redirect to the same page after updating
            except Exception as e:
                messages.error(request, f"An error occurred while updating the account: {str(e)}")
                return redirect('signup')
    return HttpResponse(template.render(context,request))

def signup_forms(request):
    template = loader.get_template('signup_form.html')
    user = AccountRegistration.objects.filter(username=request.user).values()
    context = {
        'user_role': user[0]['privilege'],
        'user_data':user[0],
    }
    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        middle_name = request.POST.get('middle_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        id_number = request.POST.get('id_number')
        contact_number = request.POST.get('contact_number')
        gender = request.POST.get('gender')
        department = request.POST.get('department')
        privilege = request.POST.get('privilege')
        status = request.POST.get('status')
        
        # Auto-generate passphrase (12 characters with letters, digits, and special chars)
        alphabet = string.ascii_letters + string.digits + '!@#$%^&*'
        password = ''.join(secrets.choice(alphabet) for i in range(12))

        print(f"Auto-generated passphrase: {password}")

        # Validate username format against privilege
        import re as _re_fmt
        _username_errors = {
            'student': ('student', r'^\d{4}-\d{5}-CL-\d+$', 'Format: XXXX-XXXXX-CL-X (e.g. 2022-00001-CL-0)'),
            'faculty': ('faculty', r'^\d{5}$', 'Format: exactly 5 digits (e.g. 00001)'),
        }
        if privilege in _username_errors:
            _, _pat, _hint = _username_errors[privilege]
            if not _re_fmt.match(_pat, username or ''):
                context['error'] = f"Invalid {privilege} username format. {_hint}"
                context['form_data'] = request.POST
                return HttpResponse(template.render(context, request))

        # Check for duplicate email
        if AccountRegistration.objects.filter(email=email).exists():
            context['error'] = f"An account with the email '{email}' already exists. Please use a different email address."
            context['form_data'] = request.POST
            return HttpResponse(template.render(context, request))
        
        # Check for duplicate ID number
        if AccountRegistration.objects.filter(id_number=id_number).exists():
            context['error'] = f"An account with the ID number '{id_number}' already exists. Please use a different ID number."
            context['form_data'] = request.POST
            return HttpResponse(template.render(context, request))
        
        # Check for duplicate username
        if AccountRegistration.objects.filter(username=username).exists():
            context['error'] = f"An account with the username '{username}' already exists. Please use a different username."
            context['form_data'] = request.POST
            return HttpResponse(template.render(context, request))

        try:
            user = AccountRegistration(
                username=username,
                first_name=first_name,
                middle_name=middle_name,
                last_name=last_name,
                email=email,
                id_number=id_number,
                contact_number=contact_number,
                gender=gender,
                department=department,
                privilege=privilege,
                status=status,
                must_change_password=True,
                password=password
            )

            user.set_password(password)
            user.save()
            sync_admin_flags(user, privilege)
            messages.success(request, 'Account created successfully!')
            # Send notification email to the newly created user with their auto-generated passphrase
            try:
                subject = 'Your ISSC Account Has Been Created'
                message = (
                    f"Welcome to ISSC, {first_name}!\n\n"
                    "Your account has been successfully created. Below are your login credentials:\n\n"
                    f"Username: {username}\n"
                    f"Passphrase: {password}\n\n"
                    "⚠️ IMPORTANT: Please keep this passphrase secure and do not share it with anyone.\n\n"
                    " It is recommended that you change your passphrase upon your first login.\n\n"
                    "You can login at: " + ('https://www.issc.study/login/') + "\n\n"
                    "If you need to reset your passphrase, please contact the administrator."
                )
                from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', settings.EMAIL_HOST_USER)

                # Log which sending account is being used (do not log the full password)
                try:
                    masked_pwd_info = f"{'*' * 4} (length={len(settings.EMAIL_HOST_PASSWORD)})" if getattr(settings, 'EMAIL_HOST_PASSWORD', None) else 'not set'
                except Exception:
                    masked_pwd_info = 'not available'
                print(f"Attempting to send email using host={getattr(settings,'EMAIL_HOST','')}, user={getattr(settings,'EMAIL_HOST_USER','')} pwd={masked_pwd_info}")

                # Create an explicit connection using current settings so credentials are clear
                connection = get_connection(
                    host=getattr(settings, 'EMAIL_HOST', None),
                    port=getattr(settings, 'EMAIL_PORT', None),
                    username=getattr(settings, 'EMAIL_HOST_USER', None),
                    password=getattr(settings, 'EMAIL_HOST_PASSWORD', None),
                    use_tls=getattr(settings, 'EMAIL_USE_TLS', False),
                )

                send_mail(subject, message, from_email, [email], connection=connection, fail_silently=False)
            except BadHeaderError:
                print('Invalid header found when sending account creation email')
            except Exception as e:
                # Log the error server-side and show a non-blocking message
                print(f"Failed sending account creation email to {email}: {e}")
            return redirect('signup')
        except Exception as e:
            context['error'] = f"An error occurred while creating the account: {str(e)}"
            context['form_data'] = request.POST
            return HttpResponse(template.render(context, request))
    return HttpResponse(template.render(context,request))

def account_details(request, username):
    """View to display and edit account details"""
    template = loader.get_template('signup_form.html')
    current_user = AccountRegistration.objects.filter(username=request.user).values()
    
    # Get the account to be edited
    account = get_object_or_404(AccountRegistration, username=username)
    
    context = {
        'user_role': current_user[0]['privilege'],
        'user_data': current_user[0],
        'account': account,
        'is_edit_mode': True,
    }
    
    if request.method == 'POST':
        if 'delete' in request.POST:
            # Handle delete action
            account.delete()
            messages.success(request, f'Account {username} has been deleted successfully!')
            return redirect('signup')
        
        if 'update' in request.POST:
            # Handle update action
            first_name = request.POST.get('first_name')
            middle_name = request.POST.get('middle_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email')
            id_number = request.POST.get('id_number')
            contact_number = request.POST.get('contact_number')
            gender = request.POST.get('gender')
            department = request.POST.get('department')
            privilege = request.POST.get('privilege')
            status = request.POST.get('status')
            
            # Check for duplicate email (excluding current account)
            if AccountRegistration.objects.filter(email=email).exclude(username=username).exists():
                context['error'] = f"An account with the email '{email}' already exists. Please use a different email address."
                context['form_data'] = request.POST
                return HttpResponse(template.render(context, request))
            
            # Check for duplicate ID number (excluding current account)
            if AccountRegistration.objects.filter(id_number=id_number).exclude(username=username).exists():
                context['error'] = f"An account with the ID number '{id_number}' already exists. Please use a different ID number."
                context['form_data'] = request.POST
                return HttpResponse(template.render(context, request))
            
            try:
                account.first_name = first_name
                account.middle_name = middle_name
                account.last_name = last_name
                account.email = email
                account.id_number = id_number
                account.contact_number = contact_number
                account.gender = gender
                account.department = department
                account.privilege = privilege
                account.status = status
                account.save()
                sync_admin_flags(account, privilege)
                messages.success(request, f'Account {username} has been updated successfully!')
                return redirect('signup')
            except Exception as e:
                context['error'] = f"An error occurred while updating the account: {str(e)}"
                context['form_data'] = request.POST
                return HttpResponse(template.render(context, request))
    
    return HttpResponse(template.render(context, request))

def logout(request):
    auth_logout(request)
    return redirect('login')


@login_required(login_url='/login/')
def profile(request):
    """Allow any logged-in user (admin/faculty/student) to manage their profile."""
    account = get_object_or_404(AccountRegistration, username=request.user.username)
    user_data = AccountRegistration.objects.filter(username=request.user.username).values().first()

    if request.method == 'POST':
        # Update profile details
        if 'profile_update' in request.POST:
            first_name = request.POST.get('first_name', '').strip()
            middle_name = request.POST.get('middle_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            contact_number = request.POST.get('contact_number', '').strip()
            gender = request.POST.get('gender', '').strip()

            # Unique email check
            if AccountRegistration.objects.filter(email=email).exclude(pk=account.pk).exists():
                messages.error(request, 'Email is already in use by another account.')
            else:
                account.first_name = first_name
                account.middle_name = middle_name
                account.last_name = last_name
                account.email = email
                account.contact_number = contact_number
                account.gender = gender
                account.save()
                messages.success(request, 'Profile updated successfully.')

        # Change password
        elif 'password_change' in request.POST:
            current_password = request.POST.get('current_password', '')
            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')

            if not account.check_password(current_password):
                messages.error(request, 'Current password is incorrect.')
            elif new_password != confirm_password:
                messages.error(request, 'New passwords do not match.')
            elif not validate_password_strength(new_password):
                messages.error(request, PASSWORD_POLICY_MESSAGE)
            else:
                account.set_password(new_password)
                account.save()
                auth_login(request, account)
                messages.success(request, 'Password changed successfully.')

        # Delete account
        elif 'delete_account' in request.POST:
            confirm_text = request.POST.get('confirm_delete', '').strip().lower()
            if confirm_text in ['delete', account.username.lower()]:
                auth_logout(request)
                account.delete()
                return redirect('login')
            else:
                messages.error(request, 'Type DELETE or your username to confirm account deletion.')

    context = {
        'user_role': account.privilege,
        'user_data': user_data,
        'account': account,
    }
    template = loader.get_template('profile.html')
    return HttpResponse(template.render(context, request))

def import_data(request):
    template = loader.get_template('import.html')
    user = AccountRegistration.objects.filter(username=request.user).values()
    context = {
        'user_role': user[0]['privilege'],
        'user_data': user[0],
    }

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            context['error'] = "Please select an Excel file."
        elif not excel_file.name.endswith(('.xlsx', '.xls')):
            context['error'] = "Please upload a valid Excel file (.xlsx or .xls)."
        else:
            try:
                df = pd.read_excel(excel_file)
                
                # Validate required columns (privilege is auto-detected from username format)
                required_columns = ['username', 'first_name', 'middle_name', 'last_name', 'email', 'contact_number', 'gender', 'department']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    context['error'] = f"Missing required columns: {', '.join(missing_columns)}"
                    return HttpResponse(template.render(context, request))
                
                success_count = 0
                error_count = 0
                errors = []
                successful_users = []  # Store successfully created users with their passwords
                
                for index, row in df.iterrows():
                    try:
                        username = str(row['username']).strip()
                        email = str(row['email']).strip()
                        first_name = str(row['first_name']).strip()
                        middle_name = str(row.get('middle_name', '')).strip()
                        last_name = str(row['last_name']).strip()
                        contact_number = str(row.get('contact_number', '')).strip()
                        gender = str(row['gender']).strip().upper()
                        department = str(row['department']).strip().upper()

                        # Auto-detect privilege from username format:
                        #   Student : XXXX-XXXXX-CL-X  (e.g. 2022-00001-CL-0)
                        #   Faculty : 00001  (exactly 5 digits)
                        #   Admin   : ADMIN001 (starts with ADMIN, case-insensitive)
                        import re
                        if re.match(r'^\d{4}-\d{5}-CL-\d+$', username):
                            privilege = 'student'
                        elif re.match(r'^\d{5}$', username):
                            privilege = 'faculty'
                        elif re.match(r'^ADMIN\d+$', username, re.IGNORECASE):
                            privilege = 'admin'
                        else:
                            error_msg = f"Row {index + 2}: Cannot determine privilege from username '{username}'. Use format XXXX-XXXXX-CL-X (student, e.g. 2022-00001-CL-0), 00001 (faculty), or ADMIN001 (admin)"
                            errors.append(error_msg)
                            error_count += 1
                            continue

                        # Validate gender
                        if gender not in ['M', 'F', 'O', 'MALE', 'FEMALE', 'OTHER']:
                            error_msg = f"Row {index + 2}: Invalid gender '{gender}'. Must be Male, Female, or Other"
                            errors.append(error_msg)
                            error_count += 1
                            continue
                        
                        # Normalize gender
                        gender_map = {'MALE': 'M', 'FEMALE': 'F', 'OTHER': 'O', 'M': 'M', 'F': 'F', 'O': 'O'}
                        gender = gender_map.get(gender, 'O')
                        
                        # Validate department
                        VALID_DEPARTMENTS = [
                            'BSIT 1-1', 'BSIT 1-2', 'BSIT 2-1', 'BSIT 2-2',
                            'BSIT 3-1', 'BSIT 3-2', 'BSIT 4-1',
                            'BSENT 1-1', 'BSENT 1-2', 'BSENT 2-1', 'BSENT 2-2',
                            'BSENT 3-1', 'BSENT 3-2', 'BSENT 4-1', 'BSENT 4-2',
                            'BTLED 1-1', 'BTLED 1-2', 'BTLED 2-1', 'BTLED 2-2',
                            'BTLED 3-1', 'BTLED 3-2', 'BTLED 4-1', 'BTLED 4-2',
                        ]
                        if department not in VALID_DEPARTMENTS:
                            error_msg = f"Row {index + 2}: Invalid department '{department}'. Must be one of: {', '.join(VALID_DEPARTMENTS)}"
                            errors.append(error_msg)
                            error_count += 1
                            continue

                        # Check for duplicate username
                        if AccountRegistration.objects.filter(username=username).exists():
                            error_msg = f"Row {index + 2}: Username '{username}' already exists"
                            errors.append(error_msg)
                            error_count += 1
                            continue
                        
                        # Check for duplicate email
                        if AccountRegistration.objects.filter(email=email).exists():
                            error_msg = f"Row {index + 2}: Email '{email}' already exists"
                            errors.append(error_msg)
                            error_count += 1
                            continue
                        
                        # Auto-generate passphrase
                        alphabet = string.ascii_letters + string.digits + '!@#$%^&*'
                        password = ''.join(secrets.choice(alphabet) for i in range(12))
                        
                        # Create user
                        user_obj = AccountRegistration(
                            username=username,
                            first_name=first_name,
                            middle_name=middle_name,
                            last_name=last_name,
                            email=email,
                            id_number=username,  # Use username as ID number
                            contact_number=contact_number,
                            gender=gender,
                            department=department,
                            privilege=privilege,
                            status='allowed',
                            must_change_password=True,
                        )
                        user_obj.set_password(password)
                        user_obj.save()
                        sync_admin_flags(user_obj, privilege)
                        
                        # Store successful user info for email
                        successful_users.append({
                            'username': username,
                            'email': email,
                            'password': password,
                            'first_name': first_name,
                            'last_name': last_name
                        })
                        
                        print(f"Saved User: {first_name} {last_name} - {username}")
                        success_count += 1
                        
                    except Exception as inner_e:
                        error_msg = f"Row {index + 2}: {str(inner_e)}"
                        errors.append(error_msg)
                        print(f"Error saving user row {index + 2}: {inner_e}")
                        error_count += 1
                
                # Send emails to all successfully created users
                if successful_users:
                    email_success = 0
                    email_failed = 0
                    
                    for user_data in successful_users:
                        try:
                            subject = 'Your ISSC Account Has Been Created'
                            message = (
                                f"Welcome to ISSC, {user_data['first_name']}!\n\n"
                                "Your account has been successfully created. Below are your login credentials:\n\n"
                                f"Username: {username}\n"
                                f"Passphrase: {password}\n\n"
                                "⚠️ IMPORTANT: Please keep this passphrase secure and do not share it with anyone.\n\n"
                                " It is recommended that you change your passphrase upon your first login.\n\n"
                                "You can login at: " + ('https://www.issc.study/login/') + "\n\n"
                                "If you need to reset your passphrase, please contact the administrator."
                            )
                            from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', settings.EMAIL_HOST_USER)
                            
                            connection = get_connection(
                                host=getattr(settings, 'EMAIL_HOST', None),
                                port=getattr(settings, 'EMAIL_PORT', None),
                                username=getattr(settings, 'EMAIL_HOST_USER', None),
                                password=getattr(settings, 'EMAIL_HOST_PASSWORD', None),
                                use_tls=getattr(settings, 'EMAIL_USE_TLS', False),
                            )
                            
                            send_mail(subject, message, from_email, [user_data['email']], connection=connection, fail_silently=False)
                            email_success += 1
                            print(f"Email sent to {user_data['email']}")
                            
                        except Exception as e:
                            email_failed += 1
                            print(f"Failed to send email to {user_data['email']}: {e}")
                    
                    if email_success > 0:
                        context['message'] = f"Successfully imported {success_count} users! Emails sent to {email_success} users."
                    if email_failed > 0:
                        if context.get('message'):
                            context['message'] += f" (Failed to send {email_failed} emails)"
                        else:
                            context['error'] = f"Imported {success_count} users but failed to send {email_failed} emails."
                
                if success_count > 0 and not context.get('message'):
                    context['message'] = f"Successfully imported {success_count} users!"
                    
                if error_count > 0:
                    error_summary = f"Failed to import {error_count} records. "
                    if errors:
                        error_summary += "Errors: " + "; ".join(errors[:5])
                        if len(errors) > 5:
                            error_summary += f"... and {len(errors) - 5} more errors."
                    
                    if context.get('error'):
                        context['error'] += " " + error_summary
                    else:
                        context['error'] = error_summary

            except Exception as e:
                context['error'] = f"Error processing Excel file: {str(e)}"
                print(f"Error processing Excel file: {str(e)}")

    return HttpResponse(template.render(context, request))


def download_template(request):
    """Generate and download Excel template for user import"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "User Import Template"
    
    # Define headers (privilege is auto-detected from username format — no column needed)
    headers = [
        'username',
        'first_name',
        'middle_name',
        'last_name',
        'email',
        'contact_number',
        'gender',
        'department',
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color='7a1818', end_color='7a1818', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add example data (no privilege column — auto-detected from username)
    example_data = [
        ['2022-00001-CL-0', 'Juan', 'Dela', 'Cruz', 'juan.delacruz@example.com', '09123456789', 'M', 'BSIT 1-1'],
        ['00001', 'Maria', 'Santos', 'Garcia', 'maria.garcia@example.com', '09987654321', 'F', 'BSENT 1-1'],
        ['ADMIN001', 'Pedro', 'Reyes', 'Lopez', 'pedro.lopez@example.com', '09456789123', 'M', 'BTLED 1-1'],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["ISSC User Import Template - Instructions"],
        [""],
        ["Column Descriptions:"],
        ["username", "Unique username for the user (must be unique)"],
        ["first_name", "User's first name (required)"],
        ["middle_name", "User's middle name (optional)"],
        ["last_name", "User's last name (required)"],
        ["email", "User's email address (must be unique, required)"],
        ["contact_number", "User's contact number (optional)"],
        ["gender", "Gender: M (Male), F (Female), or O (Other) - required"],
        ["department", "Department: e.g. BSIT 1-1, BSENT 2-2, BTLED 3-1 - required"],
        [""],
        ["Important Notes:"],
        ["• Privilege is AUTO-DETECTED from the username format — no privilege column needed"],
        ["• Student username  : 2022-00001-CL-0  (format: XXXX-XXXXX-CL-X)"],
        ["• Faculty username  : 00001  (exactly 5 digits)"],
        ["• Admin username    : ADMIN001  (starts with ADMIN followed by digits)"],
        ["• Passwords will be auto-generated and sent to each user's email"],
        ["• Username and email must be unique"],
        ["• Gender must be: M, F, or O (case insensitive)"],
        ["• Department must be one of the section codes e.g. BSIT 1-1"],
        ["• All users will have 'allowed' status by default"],
        ["• Delete the example rows before importing your actual data"],
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        if isinstance(row_data, list):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_instructions.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_num == 3 or row_num == 13:
                    cell.font = Font(bold=True)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Adjust instruction sheet column widths
    ws_instructions.column_dimensions['A'].width = 20
    ws_instructions.column_dimensions['B'].width = 60
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Create response
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ISSC_User_Import_Template.xlsx'
    
    return response


def getUser(request):
    user_type = request.GET.get('type', '').strip().lower()
    if user_type not in ['student', 'faculty', 'admin']:
        return JsonResponse({'error': 'Invalid or missing user type'}, status=400)

    import re
    if user_type == 'student':
        # Student format: XXXX-XXXXX-CL-X  (e.g. 2022-00001-CL-0)
        STUDENT_RE = re.compile(r'^\d{4}-(\d{5})-CL-\d+$')
        latest_num = -1
        for acc in AccountRegistration.objects.filter(privilege='student').values('username'):
            m = STUDENT_RE.match(acc['username'])
            if m:
                num = int(m.group(1))
                if num > latest_num:
                    latest_num = num
        next_num = latest_num + 1
        new_username = f"2022-{str(next_num).zfill(5)}-CL-0"
    elif user_type == 'faculty':
        # Faculty format: 00001, 00002, ...
        FACULTY_RE = re.compile(r'^(\d{5})$')
        latest_num = 0
        for acc in AccountRegistration.objects.filter(privilege='faculty').values('username'):
            m = FACULTY_RE.match(acc['username'])
            if m:
                num = int(m.group(1))
                if num > latest_num:
                    latest_num = num
        next_num = latest_num + 1
        new_username = str(next_num).zfill(5)
    else:
        # Admin format: admin0001, admin0002, ...
        prefix = 'admin'
        latest = (
            AccountRegistration.objects.filter(
                username__istartswith=prefix,
                privilege='admin'
            )
            .order_by('-username')
            .values('username')
            .first()
        )
        if latest and latest.get('username'):
            latest_username = latest['username']
            suffix = latest_username[len(prefix):]
            try:
                next_num = int(''.join(filter(str.isdigit, suffix))) + 1
            except (ValueError, TypeError):
                next_num = 1
        else:
            next_num = 1
        new_username = f"{prefix}{str(next_num).zfill(4)}"

    new_id = new_username
    return JsonResponse({'id_number': new_id, 'username': new_username})

# def password_reset(request):
#     if request.method == "POST":
#         email = request.POST.get("email")  # Get the email from the form
#         form = PasswordResetForm({"email": email})

#         if form.is_valid():
#             form.save(
#                 request=request,
#                 use_https=request.is_secure(),
#                 email_template_name="registration/password_reset_email.html"
#             )
#             messages.success(request, "A password reset link has been sent to your email.")
#             return redirect("password_reset_done")  # Redirect to a confirmation page

#         else:
#             messages.error(request, "Invalid email address. Please try again.")

#     else:
#         form = PasswordResetForm()

#     return render(request, "test.html")


from django.contrib.auth import views as auth_views
from django.urls import reverse_lazy
from .forms import CustomPasswordResetForm, CustomSetPasswordForm

class CustomPasswordResetView(auth_views.PasswordResetView):
    template_name = 'registration/acc_reset.html'
    form_class = CustomPasswordResetForm
    email_template_name = 'registration/acc_reset_email.html'
    success_url = reverse_lazy('password_reset_done')

class CustomPasswordResetDoneView(auth_views.PasswordResetDoneView):
    template_name = 'registration/acc_reset_done.html'

class CustomPasswordResetConfirmView(auth_views.PasswordResetConfirmView):
    template_name = 'registration/acc_reset_confirm.html'
    form_class = CustomSetPasswordForm
    success_url = reverse_lazy('password_reset_complete')

class CustomPasswordResetCompleteView(auth_views.PasswordResetCompleteView):
    template_name = 'registration/acc_reset_complete.html'


@login_required
def get_admin_contact(request):
    """Get the current admin contact number"""
    phone_number = SystemConfig.get_admin_contact()
    return JsonResponse({
        'success': True,
        'phone_number': phone_number
    })


@login_required
def save_admin_contact(request):
    """Save or update the admin contact number"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            phone_number = data.get('phone_number', '').strip()
            
            # Validate phone number
            if len(phone_number) != 11:
                return JsonResponse({
                    'success': False,
                    'message': 'Phone number must be exactly 11 digits'
                }, status=400)
            
            if not phone_number.startswith('09') or not phone_number.isdigit():
                return JsonResponse({
                    'success': False,
                    'message': 'Phone number must start with 09 and contain only digits'
                }, status=400)
            
            # Save to database
            SystemConfig.set_admin_contact(phone_number, request.user)
            
            return JsonResponse({
                'success': True,
                'message': 'Admin contact number saved successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    }, status=400)
